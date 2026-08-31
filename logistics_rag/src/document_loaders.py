"""
Logistics Document Loaders
Handles CSV/Excel shipment data, PDFs, route maps, and delivery logs.
"""

import csv
import json
from pathlib import Path
from typing import List

from langchain_core.documents import Document
from langchain_community.document_loaders import PyPDFLoader


class LogisticsDocumentLoader:
    """Loads and converts all logistics data types into LangChain Documents."""

    # ------------------------------------------------------------------
    # Shipment data (CSV / Excel)
    # ------------------------------------------------------------------

    def load_shipments(self, folder: Path) -> List[Document]:
        """Load shipment/order CSV files. Each row becomes a Document."""
        docs = []
        if not folder.exists():
            return docs

        for file in folder.glob("*.csv"):
            try:
                with open(file, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        # Build a readable text block from every field
                        text = self._shipment_row_to_text(row)
                        docs.append(Document(
                            page_content=text,
                            metadata={
                                "source": str(file),
                                "doc_type": "shipment",
                                "shipment_id": row.get("shipment_id", row.get("order_id", "")),
                            },
                        ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        # Excel support (requires openpyxl)
        for file in folder.glob("*.xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    text = self._shipment_row_to_text(row_dict)
                    docs.append(Document(
                        page_content=text,
                        metadata={"source": str(file), "doc_type": "shipment"},
                    ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        print(f"  📦 Loaded {len(docs)} shipment records from {folder}")
        return docs

    def _shipment_row_to_text(self, row: dict) -> str:
        """Convert a shipment row dict to a natural language summary."""
        parts = []
        # Priority fields first
        priority_keys = [
            "shipment_id", "order_id", "customer_name", "origin", "destination",
            "status", "carrier", "driver_name", "driver_id",
            "scheduled_delivery", "actual_delivery", "delay_minutes",
            "sla_deadline", "sla_status", "route_code", "zone",
        ]
        seen = set()
        for k in priority_keys:
            if k in row and row[k] not in (None, ""):
                parts.append(f"{k.replace('_', ' ').title()}: {row[k]}")
                seen.add(k)
        # Remaining fields
        for k, v in row.items():
            if k not in seen and v not in (None, ""):
                parts.append(f"{k.replace('_', ' ').title()}: {v}")
        return " | ".join(parts)

    # ------------------------------------------------------------------
    # SOPs (PDF)
    # ------------------------------------------------------------------

    def load_sops(self, folder: Path) -> List[Document]:
        """Load SOP PDF files using PyPDFLoader."""
        docs = []
        if not folder.exists():
            return docs

        for file in folder.glob("*.pdf"):
            try:
                loader = PyPDFLoader(str(file))
                pages = loader.load()
                for page in pages:
                    page.metadata["doc_type"] = "sop"
                docs.extend(pages)
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        print(f"  📄 Loaded {len(docs)} SOP pages from {folder}")
        return docs

    # ------------------------------------------------------------------
    # Routes / Zone maps (CSV or JSON)
    # ------------------------------------------------------------------

    def load_routes(self, folder: Path) -> List[Document]:
        """Load route and zone data."""
        docs = []
        if not folder.exists():
            return docs

        for file in folder.glob("*.csv"):
            try:
                with open(file, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)

                # Group into batches of 20 rows per document chunk
                batch_size = 20
                for i in range(0, len(rows), batch_size):
                    batch = rows[i: i + batch_size]
                    text = f"Route/Zone data from {file.name}:\n"
                    text += "\n".join(" | ".join(f"{k}: {v}" for k, v in r.items()) for r in batch)
                    docs.append(Document(
                        page_content=text,
                        metadata={"source": str(file), "doc_type": "route"},
                    ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        for file in folder.glob("*.json"):
            try:
                with open(file, encoding="utf-8") as f:
                    data = json.load(f)
                text = f"Route/Zone data from {file.name}:\n{json.dumps(data, indent=2)}"
                docs.append(Document(
                    page_content=text,
                    metadata={"source": str(file), "doc_type": "route"},
                ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        print(f"  🗺️  Loaded {len(docs)} route/zone chunks from {folder}")
        return docs

    # ------------------------------------------------------------------
    # Delivery logs (CSV / text)
    # ------------------------------------------------------------------

    def load_logs(self, folder: Path) -> List[Document]:
        """Load live delivery status logs."""
        docs = []
        if not folder.exists():
            return docs

        for file in folder.glob("*.csv"):
            try:
                with open(file, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)

                batch_size = 30
                for i in range(0, len(rows), batch_size):
                    batch = rows[i: i + batch_size]
                    text = f"Delivery log from {file.name}:\n"
                    text += "\n".join(
                        " | ".join(f"{k}: {v}" for k, v in r.items() if v not in (None, ""))
                        for r in batch
                    )
                    docs.append(Document(
                        page_content=text,
                        metadata={"source": str(file), "doc_type": "delivery_log"},
                    ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        for file in folder.glob("*.log"):
            try:
                text = file.read_text(encoding="utf-8")
                # Split large log files into 1500-char windows
                window = 1500
                for i in range(0, len(text), window):
                    chunk = text[i: i + window]
                    docs.append(Document(
                        page_content=chunk,
                        metadata={"source": str(file), "doc_type": "delivery_log"},
                    ))
            except Exception as e:
                print(f"⚠️  Could not load {file}: {e}")

        print(f"  📋 Loaded {len(docs)} log chunks from {folder}")
        return docs
